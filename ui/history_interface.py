# coding: utf-8
import csv
from datetime import datetime
from pathlib import Path

from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QWidget, QVBoxLayout, QHBoxLayout, QTableWidgetItem, QFileDialog

from qfluentwidgets import PushButton, SubtitleLabel, InfoBar, InfoBarPosition, TableWidget

try:
    from ..core.storage import load_results, clear_results, get_total_records_count
    from ..core.settings import get_settings
except ImportError:
    # Запасной импорт при запуске из каталога
    from core.storage import load_results, clear_results, get_total_records_count  # type: ignore
    from core.settings import get_settings  # type: ignore


class HistoryInterface(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent=parent)
        self.setObjectName('history-interface')
        self.settings = get_settings()

        self.vBox = QVBoxLayout(self)
        self.vBox.setContentsMargins(24, 24, 24, 24)
        self.vBox.setSpacing(12)

        self.title = SubtitleLabel('История результатов', self)
        self.title.setAlignment(Qt.AlignHCenter)

        self.table = TableWidget(self)
        self.table.setColumnCount(7)
        # Заголовки будут выставляться динамически в зависимости от единиц измерения
        self._update_headers()

        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.verticalHeader().setVisible(False)
        self.table.setAlternatingRowColors(True)

        self.buttonsRow = QHBoxLayout()
        self.refreshBtn = PushButton('Обновить', self)
        self.exportCsvBtn = PushButton('Экспорт CSV', self)
        self.exportExcelBtn = PushButton('Экспорт Excel', self)
        self.clearBtn = PushButton('Очистить', self)
        self.buttonsRow.addStretch(1)
        self.buttonsRow.addWidget(self.refreshBtn)
        self.buttonsRow.addWidget(self.exportCsvBtn)
        self.buttonsRow.addWidget(self.exportExcelBtn)
        self.buttonsRow.addWidget(self.clearBtn)
        self.buttonsRow.addStretch(1)

        self.vBox.addWidget(self.title)
        self.vBox.addLayout(self.buttonsRow)
        self.vBox.addWidget(self.table)

        self.refreshBtn.clicked.connect(self.refresh)
        self.exportCsvBtn.clicked.connect(self.export_csv)
        self.exportExcelBtn.clicked.connect(self.export_excel)
        self.clearBtn.clicked.connect(self.clear)
        # Перестраиваем заголовки/значения при изменении настроек единиц
        self.settings.changed.connect(self._on_setting_changed)

        self.refresh()

    def _info(self, text: str):
        InfoBar.success(title='Готово', content=text, orient=Qt.Horizontal, position=InfoBarPosition.TOP, parent=self)

    def _error(self, text: str):
        InfoBar.error(title='Ошибка', content=text, orient=Qt.Horizontal, position=InfoBarPosition.TOP, parent=self)

    def _units(self) -> str:
        return self.settings.get('units', 'Mbps')

    def _update_headers(self):
        units = 'MB/s' if self._units() == 'MB/s' else 'Mbps'
        self.table.setHorizontalHeaderLabels([
            'Время',
            'Ping (ms)',
            f'Download ({units})',
            f'Upload ({units})',
            'Провайдер',
            'Город',
            'Хост',
        ])

    def _on_setting_changed(self, key: str, value):
        if key == 'units':
            self._update_headers()
            self.refresh()

    def refresh(self):
        results = load_results()
        self.table.setRowCount(len(results))
        for row, r in enumerate(results):
            s = r.get('server', {})

            # конвертация скоростей
            d_bps = (r.get('download_bps', 0.0) or 0.0)
            u_bps = (r.get('upload_bps', 0.0) or 0.0)
            if self._units() == 'MB/s':
                d_val = d_bps / 8e6
                u_val = u_bps / 8e6
            else:
                d_val = d_bps / 1e6
                u_val = u_bps / 1e6

            items = [
                QTableWidgetItem(str(r.get('timestamp', ''))),
                QTableWidgetItem(f"{r.get('ping_ms', 0):.0f}"),
                QTableWidgetItem(f"{d_val:.2f}"),
                QTableWidgetItem(f"{u_val:.2f}"),
                QTableWidgetItem(str(s.get('sponsor', ''))),
                QTableWidgetItem(str(s.get('name', ''))),
                QTableWidgetItem(str(s.get('host', ''))),
            ]
            for col, item in enumerate(items):
                item.setFlags(item.flags() ^ Qt.ItemIsEditable)
                self.table.setItem(row, col, item)

        self.table.resizeColumnsToContents()

    def clear(self):
        try:
            clear_results()
            self.refresh()
            self._info('История очищена')
        except Exception as e:
            self._error(str(e))

    def export_csv(self):
        """Экспорт истории в CSV файл."""
        try:
            results = load_results()
            if not results:
                self._error('Нет данных для экспорта')
                return
            
            # Диалог сохранения файла
            default_name = f"speedtest_history_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                'Сохранить CSV',
                str(Path.home() / 'Documents' / default_name),
                'CSV Files (*.csv);;All Files (*)'
            )
            
            if not file_path:
                return
            
            # Запись в CSV
            with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                
                # Заголовки
                units = 'MB/s' if self._units() == 'MB/s' else 'Mbps'
                writer.writerow([
                    'Время',
                    'Ping (ms)',
                    f'Download ({units})',
                    f'Upload ({units})',
                    'Провайдер',
                    'Город',
                    'Страна',
                    'Хост'
                ])
                
                # Данные
                for r in results:
                    s = r.get('server', {})
                    d_bps = (r.get('download_bps', 0.0) or 0.0)
                    u_bps = (r.get('upload_bps', 0.0) or 0.0)
                    
                    if self._units() == 'MB/s':
                        d_val = d_bps / 8e6
                        u_val = u_bps / 8e6
                    else:
                        d_val = d_bps / 1e6
                        u_val = u_bps / 1e6
                    
                    writer.writerow([
                        str(r.get('timestamp', '')),
                        f"{r.get('ping_ms', 0):.0f}",
                        f"{d_val:.2f}",
                        f"{u_val:.2f}",
                        str(s.get('sponsor', '')),
                        str(s.get('name', '')),
                        str(s.get('country', '')),
                        str(s.get('host', ''))
                    ])
            
            self._info(f'Экспортировано {len(results)} записей в CSV')
        except Exception as e:
            self._error(f'Ошибка экспорта CSV: {str(e)}')

    def export_excel(self):
        """Экспорт истории в Excel файл."""
        try:
            # Проверка наличия openpyxl
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
            except ImportError:
                self._error('Для экспорта в Excel требуется установить библиотеку openpyxl')
                return
            
            results = load_results()
            if not results:
                self._error('Нет данных для экспорта')
                return
            
            # Диалог сохранения файла
            default_name = f"speedtest_history_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                'Сохранить Excel',
                str(Path.home() / 'Documents' / default_name),
                'Excel Files (*.xlsx);;All Files (*)'
            )
            
            if not file_path:
                return
            
            # Создание Excel файла
            wb = Workbook()
            ws = wb.active
            ws.title = "История тестов"
            
            # Стили для заголовков
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Заголовки
            units = 'MB/s' if self._units() == 'MB/s' else 'Mbps'
            headers = [
                'Время',
                'Ping (ms)',
                f'Download ({units})',
                f'Upload ({units})',
                'Провайдер',
                'Город',
                'Страна',
                'Хост'
            ]
            
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Данные
            for row_idx, r in enumerate(results, start=2):
                s = r.get('server', {})
                d_bps = (r.get('download_bps', 0.0) or 0.0)
                u_bps = (r.get('upload_bps', 0.0) or 0.0)
                
                if self._units() == 'MB/s':
                    d_val = d_bps / 8e6
                    u_val = u_bps / 8e6
                else:
                    d_val = d_bps / 1e6
                    u_val = u_bps / 1e6
                
                ws.cell(row=row_idx, column=1, value=str(r.get('timestamp', '')))
                ws.cell(row=row_idx, column=2, value=float(r.get('ping_ms', 0)))
                ws.cell(row=row_idx, column=3, value=round(d_val, 2))
                ws.cell(row=row_idx, column=4, value=round(u_val, 2))
                ws.cell(row=row_idx, column=5, value=str(s.get('sponsor', '')))
                ws.cell(row=row_idx, column=6, value=str(s.get('name', '')))
                ws.cell(row=row_idx, column=7, value=str(s.get('country', '')))
                ws.cell(row=row_idx, column=8, value=str(s.get('host', '')))
            
            # Автоподбор ширины колонок
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Сохранение
            wb.save(file_path)
            self._info(f'Экспортировано {len(results)} записей в Excel')
        except Exception as e:
            self._error(f'Ошибка экспорта Excel: {str(e)}')
